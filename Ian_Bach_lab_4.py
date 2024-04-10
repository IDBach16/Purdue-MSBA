# -*- coding: utf-8 -*-
"""
Created on Wed Apr 10 12:09:02 2024

@author: IBach
"""

#Import and create an excel file

import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active

sheet.title = "commission"

# Dictionary to match a commission code

commission_dict = {'A': float(0.35), 'B': float(0.25), 'C': float(0.15)}


#Before the real processing start

more_cars = 'Y'

cars_count = int(0)

adjusted_price = float(1) - float(0.0175)


#Opens the file manifest.txt from home directory folder. Please change the account name as needed.

manifestFile = open('manifest.txt','a')


#During the real processing

while more_cars == 'Y':

    #Cars details

    cars_color = input("Please enter the color of the car: ")

    cars_year_manufactured = input("Please enter the year of the car was manufactured: ")

    cars_brand = input("Please enter the manufacturer brand of the car: ")

    cars_model = input("Please enter the model of the car: ")

    cars_name = cars_color+cars_year_manufactured+cars_brand+cars_model

    #Cars retail price

    MSRP = input("Please enter the manufaturer's suggested retail price: ")

    #Cars cost

    cars_cost = input("Please input the cost of the car: ")

    #The code of the type of car (0 = domestic or 1 = import) for each car

    cars_code = input("Please enter 0 for domestic, and 1 for import. ")

    #The code of the expected commission rate

    commission_code = input("Please input the code of the expected commission rate --> in capital letter: ")

    #Contents of new record to be added to the file to demonstrate writing to a file

    x = '\n'+str(cars_name)+' '+str(MSRP)+' '+str(cars_cost)+' '+str(cars_code)+' '+str(commission_code)

    manifestFile.write(x)

    #To stop or check-in more cars

    cars_count = cars_count + int(1)

    more_cars = input("To check in more cars, enter 'Y', otherwise, type anything.")


#Closing the file after writing to it

manifestFile.close()

#Opening the file and reading the data from it

manifestFile = open('manifest.txt')

carsContent = manifestFile.read()

#Establishing a list to separate out each line and breaking each line on the end line

carsList = []

carsList = carsContent.split('\n')


#CPrinting the contentsof the file as a list of items and then closing the file

print(carsList)

manifestFile.close()

#Processing car name will go in Column A and potential commission will go in Column B

for i in range(len(carsList)):

    breakup = carsList[i].split( ) #Splits up each line into 5 fields

    carsName = breakup[0] if len(breakup) > 0 else "0" #Places first item in line

    carsMSRP = breakup[1] if len(breakup) > 1 else "0" #Places second item in line

    carsCost = breakup[2] if len(breakup) > 2 else "0" #Places third item in line

    carsCode = breakup[3] if len(breakup) > 3 else "0" #Places fourth item in line

    commissionCode = breakup[4] if len(breakup) > 4 else "0" #Places fifth item in line

    #Potential gross profit

    potential_gross_profit = float(carsMSRP) - float(carsCost)

    #The code of the type of car (0 = domestic or 1 = import) for each car

    if carsCode == '0': #Domestic

        adjusted_gross_profit = float(potential_gross_profit)

    else: #Essentially user has to input 1 to get Import calculation

        adjusted_gross_profit = float(potential_gross_profit) * float(adjusted_price)


    #The expected commission amount

    commission_rate = commission_dict.get(commissionCode)
   

    if commission_rate is not None:

        commission_potential = adjusted_gross_profit*float(commission_rate)


        #Adjusts the row because lists start at 0, but spreadsheets start at 1

        r = i + 1

        #Car name will go in Column A and potential commission will go in Column B

        sheet.cell(row = r, column = 1).value = carsName

        sheet.cell(row = r, column = 2).value = commission_potential


    else:

        break

   

#Save the workbook

wb.save('commission.xlsx')